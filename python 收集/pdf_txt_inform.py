# In[11]:


import sys
import importlib
importlib.reload(sys)

from pdfminer.pdfparser import PDFParser,PDFDocument
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.converter import PDFPageAggregator
from pdfminer.layout import LTTextBoxHorizontal,LAParams
from pdfminer.pdfinterp import PDFTextExtractionNotAllowed

import re


# In[12]:


#解析pdf 文本，保存到txt文件中
def parse(path1,path2):
    fp = open(path1, 'rb') # 以二进制读模式打开
    #用文件对象来创建一个pdf文档分析器
    praser = PDFParser(fp)
    # 创建一个PDF文档
    doc = PDFDocument()
    # 连接分析器 与文档对象
    praser.set_document(doc)
    doc.set_parser(praser)

    # 提供初始化密码
    # 如果没有密码 就创建一个空的字符串
    doc.initialize()

    # 检测文档是否提供txt转换，不提供就忽略
    if not doc.is_extractable:
        raise PDFTextExtractionNotAllowed
    else:
        # 创建PDf 资源管理器 来管理共享资源
        rsrcmgr = PDFResourceManager()
        # 创建一个PDF设备对象
        laparams = LAParams()
        device = PDFPageAggregator(rsrcmgr, laparams=laparams)
        # 创建一个PDF解释器对象
        interpreter = PDFPageInterpreter(rsrcmgr, device)

        # 循环遍历列表，每次处理一个page的内容
        for page in doc.get_pages(): # doc.get_pages() 获取page列表
            interpreter.process_page(page)
            # 接受该页面的LTPage对象
            layout = device.get_result()
            # 这里layout是一个LTPage对象 里面存放着 这个page解析出的各种对象 一般包括LTTextBox, LTFigure, LTImage, LTTextBoxHorizontal 等等 想要获取文本就获得对象的text属性，
            for x in layout:
                if (isinstance(x, LTTextBoxHorizontal)):
                    with open(path2, 'a',encoding='utf-8') as f:
                        results = x.get_text()
                        f.write(results + '\n')


# In[48]:


import re

def get_date(path):
    with open(path,'r',encoding='utf-8') as f:
        txt = f.read()
    
    txt = txt.replace("\n","")
    
    
    try:
        name = txt.split(" ")[0]
        #申请代码1
        a1 = txt.split(" ")[1]  
        #申请代码2
        Ra2 = re.compile(r"NSFC- L\d{2}:(.*?): .+?Abstract\(limited to 500 words\)")
        a2 = Ra2.findall(txt)[0]
    
        #英文摘要
        Rlong = re.compile(r"Abstract\(limited to 500 words\)(.*?)Keywords\(limited to 5 keywords,seperated by;\)")
        long = Rlong.findall(txt)[0]
        #英文关键词
        Rshort = re.compile(r"Keywords\(limited to 5 keywords,seperated by;\):(.*?)1 2 ")
        short = Rshort.findall(txt)[0]
        #各作者的日期
        t = re.search(r"Keywords\(limited to 5 keywords,seperated by;\).*?@",txt).group(0)
        ls=re.findall('\d{4}\.\d{2}',t)
        
        return name,a1,a2,long,short,ls
    
    except:
        name = txt.split(" ")[0]
        print("error_{}".format(name))


# In[49]:


from openpyxl import load_workbook

#列表写入excel文件
            #填写要固定的列
def write(index,ws,lt):
    for i in range(len(lt)):
        ws.cell(2+i,index).value = lt[i]
        print(i,end=' ')
            #选择填入某行的单元格
                                


# ### 操作输入区
#====================================================================================================#
# In[51]:																							 #	
																									 #
																									 #
																									 #
Path1 = "2010 ({}).pdf"						                                                         #
Path2 = "2010 ({}).txt"					                                                             #
																									 #
#====================================================================================================#
d = dict()
lt_name = []
lt_a1 = []
lt_a2 = []
lt_long = []
lt_short = []
#******************************************************#
for i in range(1,29):								   #
#******************************************************#
    try:
        path1 = Path1.format(i)
        path2 = Path2.format(i)
        parse(path1,path2)
        name,a1,a2,long,short,ls = get_date(path2)
    
        lt_name.append(name)
        lt_a1.append(a1)
        lt_a2.append(a2)
        lt_long.append(long)
        lt_short.append(short)
        d[name] = ls
    except:
        print("error_{}".format(path1))


# In[52]:


#处理文件
#================================================================================#
path = 'code.xlsx'  #保存目录                                                    #
#================================================================================#
wb = load_workbook(path)
ws1 = wb["Sheet1"]   #工作表名字
ws2 = wb["Sheet2"]

#填写sheet1
write(1,ws1,lt_name)
write(3,ws1,lt_a1)
write(4,ws1,lt_a2)
write(8,ws1,lt_long)
write(9,ws1,lt_short)

print("write_Sheet1_ok")


# In[53]:



#填写sheet2
def write_index(index,ws,lt):
    for i in range(len(lt)):
        ws.cell(2+i,index).value = lt[i]
    print("write_lt_ok")
            #选择填入某行的单元格
            
def write_dic(dic,ws):
    key_lt = []
    value_lt = []
    for key in dic:
        value = dic[key]
        key_lt.extend([key for i in range(len(value))])
        value_lt.extend(value)
    
    write_index(1,ws,key_lt)
    write_index(4,ws,value_lt)
    
    
    print("write_dic_ok!")
write_dic(d,ws2)
print("write_Sheet2_ok")
wb.save(path)
